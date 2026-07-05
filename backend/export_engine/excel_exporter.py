import pandas as pd
import numpy as np
import os
from typing import Dict, Any
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelExporter:
    def __init__(self):
        self.export_dir = "exports"
        os.makedirs(self.export_dir, exist_ok=True)
    
    def export_analysis_results(self, empirical_results: Dict[str, Any], project_name: str) -> str:
        """
        导出分析结果为Excel文件（多工作表）
        """
        try:
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            excel_path = os.path.join(self.export_dir, f"实证分析结果_{project_name}_{timestamp}.xlsx")
            
            # 创建工作簿
            wb = Workbook()
            
            # 定义样式
            header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            data_font = Font(name='微软雅黑', size=10)
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 1. 描述性统计
            if "descriptive" in empirical_results:
                ws = wb.active
                ws.title = "描述性统计"
                self._add_descriptive_sheet(ws, empirical_results["descriptive"], 
                                          header_font, header_fill, data_font, center_align, thin_border)
            
            # 2. 相关性分析
            if "correlation" in empirical_results:
                ws = wb.create_sheet("相关性分析")
                self._add_correlation_sheet(ws, empirical_results["correlation"],
                                          header_font, header_fill, data_font, center_align, thin_border)
            
            # 3. VIF检验
            if "vif" in empirical_results:
                ws = wb.create_sheet("VIF检验")
                self._add_vif_sheet(ws, empirical_results["vif"],
                                  header_font, header_fill, data_font, center_align, thin_border)
            
            # 4. 回归结果
            if "regression" in empirical_results:
                ws = wb.create_sheet("回归结果")
                self._add_regression_sheet(ws, empirical_results["regression"],
                                         header_font, header_fill, data_font, center_align, thin_border)
            
            # 5. 稳健性检验
            if "robustness" in empirical_results:
                ws = wb.create_sheet("稳健性检验")
                self._add_robustness_sheet(ws, empirical_results["robustness"],
                                         header_font, header_fill, data_font, center_align, thin_border)
            
            # 保存文件
            wb.save(excel_path)
            
            logger.info(f"Excel导出成功: {excel_path}")
            return excel_path
            
        except Exception as e:
            logger.error(f"Excel导出失败: {str(e)}")
            raise
    
    def _add_descriptive_sheet(self, ws, descriptive_results, header_font, header_fill, data_font, center_align, thin_border):
        """添加描述性统计工作表"""
        variables = descriptive_results.get("variables", {})
        
        # 标题行
        headers = ['变量', '样本量', '均值', '标准差', '最小值', '中位数', '最大值']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # 数据行
        row_num = 2
        for var_name, stats in variables.items():
            ws.cell(row=row_num, column=1, value=var_name).font = data_font
            ws.cell(row=row_num, column=2, value=int(stats.get('count', 0))).font = data_font
            ws.cell(row=row_num, column=3, value=round(stats.get('mean', 0), 4)).font = data_font
            ws.cell(row=row_num, column=4, value=round(stats.get('std', 0), 4)).font = data_font
            ws.cell(row=row_num, column=5, value=round(stats.get('min', 0), 4)).font = data_font
            ws.cell(row=row_num, column=6, value=round(stats.get('median', 0), 4)).font = data_font
            ws.cell(row=row_num, column=7, value=round(stats.get('max', 0), 4)).font = data_font
            
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).alignment = center_align
                ws.cell(row=row_num, column=col).border = thin_border
            
            row_num += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 15
    
    def _add_correlation_sheet(self, ws, correlation_results, header_font, header_fill, data_font, center_align, thin_border):
        """添加相关性分析工作表"""
        pearson = correlation_results.get("pearson", {})
        p_values = correlation_results.get("p_values", {})
        
        if not pearson:
            return
        
        variables = list(pearson.keys())
        
        # 标题行
        ws.cell(row=1, column=1, value="Pearson相关系数矩阵").font = Font(name='微软雅黑', bold=True, size=12)
        ws.merge_cells('A1:G1')
        
        row_num = 3
        ws.cell(row=row_num, column=1, value="变量").font = header_font
        ws.cell(row=row_num, column=1).fill = header_fill
        ws.cell(row=row_num, column=1).alignment = center_align
        ws.cell(row=row_num, column=1).border = thin_border
        
        for col, var in enumerate(variables, 2):
            cell = ws.cell(row=row_num, column=col, value=var)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # 数据行
        row_num = 4
        for var1 in variables:
            ws.cell(row=row_num, column=1, value=var1).font = data_font
            ws.cell(row=row_num, column=1).alignment = center_align
            ws.cell(row=row_num, column=1).border = thin_border
            
            for col, var2 in enumerate(variables, 2):
                corr_value = pearson[var1].get(var2, 0)
                p_value = p_values.get(var1, {}).get(var2, 1)
                
                # 添加显著性标记
                display_value = f"{corr_value:.4f}"
                if p_value < 0.01:
                    display_value += "***"
                elif p_value < 0.05:
                    display_value += "**"
                elif p_value < 0.1:
                    display_value += "*"
                
                cell = ws.cell(row=row_num, column=col, value=display_value)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border
            
            row_num += 1
        
        # 添加显著性说明
        row_num += 1
        ws.cell(row=row_num, column=1, value="显著性水平: *** p<0.01, ** p<0.05, * p<0.1").font = Font(name='微软雅黑', size=9, italic=True)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        for col in range(2, len(variables) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _add_vif_sheet(self, ws, vif_results, header_font, header_fill, data_font, center_align, thin_border):
        """添加VIF检验工作表"""
        vif_data = vif_results.get("vif_values", {})
        
        # 标题行
        headers = ['变量', 'VIF值', '判断']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # 数据行
        row_num = 2
        for var_name, vif_value in vif_data.items():
            if vif_value > 10:
                judgment = "严重共线性"
            elif vif_value > 5:
                judgment = "中度共线性"
            else:
                judgment = "无共线性问题"
            
            ws.cell(row=row_num, column=1, value=var_name).font = data_font
            ws.cell(row=row_num, column=2, value=round(vif_value, 4)).font = data_font
            ws.cell(row=row_num, column=3, value=judgment).font = data_font
            
            for col in range(1, 4):
                ws.cell(row=row_num, column=col).alignment = center_align
                ws.cell(row=row_num, column=col).border = thin_border
            
            row_num += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
    
    def _add_regression_sheet(self, ws, regression_results, header_font, header_fill, data_font, center_align, thin_border):
        """添加回归结果工作表"""
        row_num = 1
        
        for model_name, model_result in regression_results.items():
            if not isinstance(model_result, dict):
                continue
            
            # 模型标题
            ws.cell(row=row_num, column=1, value=f"{model_result.get('model_name', model_name)} 回归结果").font = Font(name='微软雅黑', bold=True, size=12)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
            row_num += 2
            
            # 表头
            headers = ['变量', '系数', '标准误', 't值', 'P值', '显著性']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            row_num += 1
            
            # 回归系数
            coefficients = model_result.get("coefficients", {})
            for var_name, coef_data in coefficients.items():
                coef = coef_data.get('coef', 0)
                std_err = coef_data.get('std_err', 0)
                t_value = coef_data.get('t_value', 0)
                p_value = coef_data.get('p_value', 1)
                
                # 显著性标记
                if p_value < 0.01:
                    significance = '***'
                elif p_value < 0.05:
                    significance = '**'
                elif p_value < 0.1:
                    significance = '*'
                else:
                    significance = ''
                
                ws.cell(row=row_num, column=1, value=var_name).font = data_font
                ws.cell(row=row_num, column=2, value=round(coef, 4)).font = data_font
                ws.cell(row=row_num, column=3, value=round(std_err, 4)).font = data_font
                ws.cell(row=row_num, column=4, value=round(t_value, 4)).font = data_font
                ws.cell(row=row_num, column=5, value=round(p_value, 4)).font = data_font
                ws.cell(row=row_num, column=6, value=significance).font = data_font
                
                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).alignment = center_align
                    ws.cell(row=row_num, column=col).border = thin_border
                
                row_num += 1
            
            # 模型统计量
            row_num += 1
            stats = model_result.get("model_stats", {})
            ws.cell(row=row_num, column=1, value="模型统计量:").font = Font(name='微软雅黑', bold=True, size=10)
            row_num += 1
            ws.cell(row=row_num, column=1, value=f"样本量: {stats.get('n_obs', 'N/A')}").font = data_font
            row_num += 1
            ws.cell(row=row_num, column=1, value=f"R²: {stats.get('r_squared', 'N/A'):.4f}").font = data_font
            row_num += 1
            ws.cell(row=row_num, column=1, value=f"调整R²: {stats.get('adj_r_squared', 'N/A'):.4f}").font = data_font
            row_num += 1
            ws.cell(row=row_num, column=1, value=f"F值: {stats.get('f_statistic', 'N/A'):.4f}").font = data_font
            row_num += 2
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
    
    def _add_robustness_sheet(self, ws, robustness_results, header_font, header_fill, data_font, center_align, thin_border):
        """添加稳健性检验工作表"""
        row_num = 1
        
        if not isinstance(robustness_results, dict):
            ws.cell(row=row_num, column=1, value="无稳健性检验数据").font = data_font
            return
        
        for test_name, test_result in robustness_results.items():
            # 检验标题
            ws.cell(row=row_num, column=1, value=test_result.get("test_name", test_name)).font = Font(name='微软雅黑', bold=True, size=12)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
            row_num += 2
            
            # 检验说明
            description = test_result.get("description", "")
            if description:
                ws.cell(row=row_num, column=1, value=description).font = Font(name='微软雅黑', size=10)
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
                row_num += 2
            
            # 结果表格
            if "results" in test_result:
                results = test_result["results"]
                if isinstance(results, dict):
                    # 表头
                    ws.cell(row=row_num, column=1, value="指标").font = header_font
                    ws.cell(row=row_num, column=1).fill = header_fill
                    ws.cell(row=row_num, column=1).alignment = center_align
                    ws.cell(row=row_num, column=1).border = thin_border
                    
                    ws.cell(row=row_num, column=2, value="结果").font = header_font
                    ws.cell(row=row_num, column=2).fill = header_fill
                    ws.cell(row=row_num, column=2).alignment = center_align
                    ws.cell(row=row_num, column=2).border = thin_border
                    row_num += 1
                    
                    # 数据行
                    for key, value in results.items():
                        ws.cell(row=row_num, column=1, value=key).font = data_font
                        ws.cell(row=row_num, column=2, value=str(value)).font = data_font
                        ws.cell(row=row_num, column=1).alignment = center_align
                        ws.cell(row=row_num, column=1).border = thin_border
                        ws.cell(row=row_num, column=2).alignment = center_align
                        ws.cell(row=row_num, column=2).border = thin_border
                        row_num += 1
            
            row_num += 2
        
        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
